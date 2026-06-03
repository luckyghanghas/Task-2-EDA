import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np

def build_beautiful_mockup():
    print("🎨 Formatting and building a premium sales dashboard mockup...")
    
    # 1. Read cleaned data
    df = pd.read_csv('data/sales_transactions_cleaned.csv')
    df['order_date'] = pd.to_datetime(df['order_date'])
    df['year_month'] = df['order_date'].dt.to_period('M').astype(str)
    
    # Calculate exact metrics
    total_revenue = df['revenue'].sum()
    total_orders = len(df)
    avg_order_value = df['revenue'].mean()
    total_profit = df['gross_profit'].sum()
    profit_margin = total_profit / total_revenue
    avg_rating = df['customer_rating'].mean()
    avg_delivery = df['delivery_days'].mean()
    return_rate = df['returned_flag'].mean()

    # Create Workbook
    wb = openpyxl.Workbook()
    
    # Setup styles
    font_family = "Segoe UI"
    
    # Palette: Indigo Theme (Sleek corporate dashboard)
    fill_header = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid") # Dark Charcoal
    fill_accent = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo Accent
    fill_card = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid") # Off-white card background
    fill_zebra = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid") # Zebra light gray
    
    font_title = Font(name=font_family, size=18, bold=True, color="FFFFFF")
    font_section = Font(name=font_family, size=13, bold=True, color="1F2937")
    font_header = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    font_data = Font(name=font_family, size=10)
    font_bold = Font(name=font_family, size=10, bold=True)
    font_kpi_num = Font(name=font_family, size=16, bold=True, color="4F46E5")
    font_kpi_label = Font(name=font_family, size=9, color="6B7280", bold=True)
    
    thin_border_side = Side(border_style="thin", color="E5E7EB")
    double_border_side = Side(border_style="double", color="374151")
    thick_bottom_side = Side(border_style="medium", color="4F46E5")
    
    border_data = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_kpi = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thick_bottom_side)
    border_total = Border(top=thin_border_side, bottom=double_border_side)
    
    # ----------------------------------------------------
    # TAB 1: EXECUTIVE DASHBOARD
    # ----------------------------------------------------
    ws_dash = wb.active
    ws_dash.title = "Executive Dashboard"
    ws_dash.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws_dash.merge_cells("A1:K2")
    ws_dash["A1"] = "SALES PERFORMANCE EXECUTIVE DASHBOARD"
    ws_dash["A1"].font = font_title
    ws_dash["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for r in range(1, 3):
        for c in range(1, 12):
            ws_dash.cell(row=r, column=c).fill = fill_header
            
    # Subtitle / Description
    ws_dash["A3"] = "Interactive executive static mockup representing the cleaned retail transactions database."
    ws_dash["A3"].font = Font(name=font_family, size=10, italic=True, color="6B7280")
    ws_dash.row_dimensions[1].height = 25
    ws_dash.row_dimensions[2].height = 20
    ws_dash.row_dimensions[3].height = 20
    
    # 5 KPI Cards (Rows 5 to 7)
    kpis = [
        {"col_start": 1, "col_end": 2, "label": "TOTAL REVENUE", "value": total_revenue, "format": "$#,##0.00"},
        {"col_start": 3, "col_end": 4, "label": "TOTAL ORDERS", "value": total_orders, "format": "#,##0"},
        {"col_start": 5, "col_end": 6, "label": "AVERAGE ORDER VALUE", "value": avg_order_value, "format": "$#,##0.00"},
        {"col_start": 7, "col_end": 8, "label": "GROSS MARGIN", "value": profit_margin, "format": "0.00%"},
        {"col_start": 9, "col_end": 10, "label": "RETURN RATE", "value": return_rate, "format": "0.00%"}
    ]
    
    for kpi in kpis:
        cs = kpi["col_start"]
        ce = kpi["col_end"]
        
        # Merge labels (Row 5)
        ws_dash.merge_cells(start_row=5, start_column=cs, end_row=5, end_column=ce)
        cell_lbl = ws_dash.cell(row=5, column=cs)
        cell_lbl.value = kpi["label"]
        cell_lbl.font = font_kpi_label
        cell_lbl.alignment = Alignment(horizontal="center", vertical="center")
        
        # Merge numbers (Row 6)
        ws_dash.merge_cells(start_row=6, start_column=cs, end_row=6, end_column=ce)
        cell_val = ws_dash.cell(row=6, column=cs)
        cell_val.value = kpi["value"]
        cell_val.font = font_kpi_num
        cell_val.number_format = kpi["format"]
        cell_val.alignment = Alignment(horizontal="center", vertical="center")
        
        # Apply style to card blocks
        for r in [5, 6]:
            for c in range(cs, ce + 1):
                cell = ws_dash.cell(row=r, column=c)
                cell.fill = fill_card
                cell.border = border_kpi
                
    ws_dash.row_dimensions[5].height = 18
    ws_dash.row_dimensions[6].height = 28
    
    # Sections Row 9: Top Products by Revenue
    ws_dash["A9"] = "TOP PRODUCTS BY REVENUE"
    ws_dash["A9"].font = font_section
    ws_dash["G9"] = "REVENUE BY SALES CHANNEL"
    ws_dash["G9"].font = font_section
    
    # Headers
    headers_prod = ["Product Name", "Revenue ($)", "Orders", "Avg Price ($)", "Return Rate (%)"]
    for i, h in enumerate(headers_prod):
        cell = ws_dash.cell(row=10, column=i+1)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_accent
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    headers_chan = ["Sales Channel", "Revenue ($)", "Orders", "Gross Profit ($)", "Profit Margin %"]
    for i, h in enumerate(headers_chan):
        cell = ws_dash.cell(row=10, column=i+7)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_accent
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    ws_dash.row_dimensions[10].height = 22
    
    # Get Top Product stats
    top_p = df.groupby('product_name').agg({
        'revenue': 'sum',
        'order_id': 'count',
        'unit_price': 'mean',
        'returned_flag': 'mean'
    }).sort_values('revenue', ascending=False)
    
    # Write Top Product Data
    for idx, (pname, row_data) in enumerate(top_p.iterrows()):
        r_idx = 11 + idx
        ws_dash.row_dimensions[r_idx].height = 18
        
        c1 = ws_dash.cell(row=r_idx, column=1, value=pname)
        c2 = ws_dash.cell(row=r_idx, column=2, value=row_data['revenue'])
        c3 = ws_dash.cell(row=r_idx, column=3, value=row_data['order_id'])
        c4 = ws_dash.cell(row=r_idx, column=4, value=row_data['unit_price'])
        c5 = ws_dash.cell(row=r_idx, column=5, value=row_data['returned_flag'])
        
        c2.number_format = "$#,##0.00"
        c3.number_format = "#,##0"
        c4.number_format = "$#,##0.00"
        c5.number_format = "0.00%"
        
        c1.alignment = Alignment(horizontal="left", vertical="center")
        for cell in [c2, c3, c4, c5]:
            cell.alignment = Alignment(horizontal="right", vertical="center")
            
        for cell in [c1, c2, c3, c4, c5]:
            cell.font = font_data
            cell.border = border_data
            if idx % 2 == 1:
                cell.fill = fill_zebra
                
    # Get Channel Stats
    top_c = df.groupby('sales_channel').agg({
        'revenue': 'sum',
        'order_id': 'count',
        'gross_profit': 'sum'
    }).sort_values('revenue', ascending=False)
    
    # Write Channel Data
    for idx, (cname, row_data) in enumerate(top_c.iterrows()):
        r_idx = 11 + idx
        c1 = ws_dash.cell(row=r_idx, column=7, value=cname)
        c2 = ws_dash.cell(row=r_idx, column=8, value=row_data['revenue'])
        c3 = ws_dash.cell(row=r_idx, column=9, value=row_data['order_id'])
        c4 = ws_dash.cell(row=r_idx, column=10, value=row_data['gross_profit'])
        c5 = ws_dash.cell(row=r_idx, column=11, value=row_data['gross_profit'] / row_data['revenue'])
        
        c2.number_format = "$#,##0.00"
        c3.number_format = "#,##0"
        c4.number_format = "$#,##0.00"
        c5.number_format = "0.00%"
        
        c1.alignment = Alignment(horizontal="left", vertical="center")
        for cell in [c2, c3, c4, c5]:
            cell.alignment = Alignment(horizontal="right", vertical="center")
            
        for cell in [c1, c2, c3, c4, c5]:
            cell.font = font_data
            cell.border = border_data
            if idx % 2 == 1:
                cell.fill = fill_zebra
                
    # Add Section Row 20: Performance by Region & Campaign ROI
    ws_dash["A21"] = "PERFORMANCE BY REGION"
    ws_dash["A21"].font = font_section
    ws_dash["G21"] = "MARKETING CAMPAIGN SUMMARY"
    ws_dash["G21"].font = font_section
    
    # Headers Row 22
    headers_reg = ["Region", "Revenue ($)", "Orders", "Gross Profit ($)", "Return Rate (%)"]
    for i, h in enumerate(headers_reg):
        cell = ws_dash.cell(row=22, column=i+1)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_accent
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    headers_camp = ["Campaign", "Revenue ($)", "Orders", "Gross Profit ($)", "Profit Margin %"]
    for i, h in enumerate(headers_camp):
        cell = ws_dash.cell(row=22, column=i+7)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_accent
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    ws_dash.row_dimensions[22].height = 22
    
    # Get Region stats
    top_r = df.groupby('region').agg({
        'revenue': 'sum',
        'order_id': 'count',
        'gross_profit': 'sum',
        'returned_flag': 'mean'
    }).sort_values('revenue', ascending=False)
    
    # Write Region Data
    for idx, (rname, row_data) in enumerate(top_r.iterrows()):
        r_idx = 23 + idx
        ws_dash.row_dimensions[r_idx].height = 18
        
        c1 = ws_dash.cell(row=r_idx, column=1, value=rname)
        c2 = ws_dash.cell(row=r_idx, column=2, value=row_data['revenue'])
        c3 = ws_dash.cell(row=r_idx, column=3, value=row_data['order_id'])
        c4 = ws_dash.cell(row=r_idx, column=4, value=row_data['gross_profit'])
        c5 = ws_dash.cell(row=r_idx, column=5, value=row_data['returned_flag'])
        
        c2.number_format = "$#,##0.00"
        c3.number_format = "#,##0"
        c4.number_format = "$#,##0.00"
        c5.number_format = "0.00%"
        
        c1.alignment = Alignment(horizontal="left", vertical="center")
        for cell in [c2, c3, c4, c5]:
            cell.alignment = Alignment(horizontal="right", vertical="center")
            
        for cell in [c1, c2, c3, c4, c5]:
            cell.font = font_data
            cell.border = border_data
            if idx % 2 == 1:
                cell.fill = fill_zebra
                
    # Get Campaign Stats
    top_cam = df.groupby('campaign').agg({
        'revenue': 'sum',
        'order_id': 'count',
        'gross_profit': 'sum'
    }).sort_values('revenue', ascending=False)
    
    # Write Campaign Data
    for idx, (camname, row_data) in enumerate(top_cam.iterrows()):
        r_idx = 23 + idx
        c1 = ws_dash.cell(row=r_idx, column=7, value=camname)
        c2 = ws_dash.cell(row=r_idx, column=8, value=row_data['revenue'])
        c3 = ws_dash.cell(row=r_idx, column=9, value=row_data['order_id'])
        c4 = ws_dash.cell(row=r_idx, column=10, value=row_data['gross_profit'])
        c5 = ws_dash.cell(row=r_idx, column=11, value=row_data['gross_profit'] / row_data['revenue'])
        
        c2.number_format = "$#,##0.00"
        c3.number_format = "#,##0"
        c4.number_format = "$#,##0.00"
        c5.number_format = "0.00%"
        
        c1.alignment = Alignment(horizontal="left", vertical="center")
        for cell in [c2, c3, c4, c5]:
            cell.alignment = Alignment(horizontal="right", vertical="center")
            
        for cell in [c1, c2, c3, c4, c5]:
            cell.font = font_data
            cell.border = border_data
            if idx % 2 == 1:
                cell.fill = fill_zebra

    # Column Width Auto-Adjustment for Dashboard
    for col in ws_dash.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        # Avoid title row length distortion
        for cell in col:
            if cell.row > 2 and cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws_dash.column_dimensions[col_letter].width = max(max_len + 4, 12)
    ws_dash.column_dimensions["A"].width = 24
    ws_dash.column_dimensions["G"].width = 22
    ws_dash.column_dimensions["F"].width = 4
    
    # ----------------------------------------------------
    # TAB 2: MONTHLY TREND
    # ----------------------------------------------------
    ws_trend = wb.create_sheet(title="Monthly Trend Analysis")
    ws_trend.views.sheetView[0].showGridLines = True
    
    ws_trend["A1"] = "MONTHLY SALES AND PROFIT TREND"
    ws_trend["A1"].font = Font(name=font_family, size=15, bold=True, color="1F2937")
    ws_trend.row_dimensions[1].height = 24
    
    headers_trend = ["Order Month", "Total Revenue ($)", "Order Count", "Total Cost ($)", "Gross Profit ($)", "Profit Margin %"]
    for i, h in enumerate(headers_trend):
        cell = ws_trend.cell(row=3, column=i+1)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_trend.row_dimensions[3].height = 22
    
    monthly_data = df.groupby('year_month').agg({
        'revenue': 'sum',
        'order_id': 'count',
        'cost': 'sum',
        'gross_profit': 'sum'
    }).sort_index()
    
    for idx, (mname, row_data) in enumerate(monthly_data.iterrows()):
        r_idx = 4 + idx
        ws_trend.row_dimensions[r_idx].height = 18
        
        c1 = ws_trend.cell(row=r_idx, column=1, value=str(mname))
        c2 = ws_trend.cell(row=r_idx, column=2, value=row_data['revenue'])
        c3 = ws_trend.cell(row=r_idx, column=3, value=row_data['order_id'])
        c4 = ws_trend.cell(row=r_idx, column=4, value=row_data['cost'])
        c5 = ws_trend.cell(row=r_idx, column=5, value=row_data['gross_profit'])
        c6 = ws_trend.cell(row=r_idx, column=6, value=row_data['gross_profit'] / row_data['revenue'])
        
        c2.number_format = "$#,##0.00"
        c3.number_format = "#,##0"
        c4.number_format = "$#,##0.00"
        c5.number_format = "$#,##0.00"
        c6.number_format = "0.00%"
        
        c1.alignment = Alignment(horizontal="center", vertical="center")
        for cell in [c2, c3, c4, c5, c6]:
            cell.alignment = Alignment(horizontal="right", vertical="center")
            
        for cell in [c1, c2, c3, c4, c5, c6]:
            cell.font = font_data
            cell.border = border_data
            if idx % 2 == 1:
                cell.fill = fill_zebra
                
    # Add Totals Row
    tot_row = 4 + len(monthly_data)
    ws_trend.cell(row=tot_row, column=1, value="Total").font = font_bold
    ws_trend.cell(row=tot_row, column=1).alignment = Alignment(horizontal="center")
    
    t_rev = ws_trend.cell(row=tot_row, column=2, value=f"=SUM(B4:B{tot_row-1})")
    t_cnt = ws_trend.cell(row=tot_row, column=3, value=f"=SUM(C4:C{tot_row-1})")
    t_cost = ws_trend.cell(row=tot_row, column=4, value=f"=SUM(D4:D{tot_row-1})")
    t_prof = ws_trend.cell(row=tot_row, column=5, value=f"=SUM(E4:E{tot_row-1})")
    t_marg = ws_trend.cell(row=tot_row, column=6, value=f"=E{tot_row}/B{tot_row}")
    
    t_rev.number_format = "$#,##0.00"
    t_cnt.number_format = "#,##0"
    t_cost.number_format = "$#,##0.00"
    t_prof.number_format = "$#,##0.00"
    t_marg.number_format = "0.00%"
    
    for c in range(1, 7):
        cell = ws_trend.cell(row=tot_row, column=c)
        cell.font = font_bold
        cell.border = border_total
        
    for col in ws_trend.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 1 and cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws_trend.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    # ----------------------------------------------------
    # TAB 3: BREAKDOWNS (Customer & Product Category detail)
    # ----------------------------------------------------
    ws_break = wb.create_sheet(title="Category & Segment Breakdowns")
    ws_break.views.sheetView[0].showGridLines = True
    
    ws_break["A1"] = "SALES BY PRODUCT CATEGORY"
    ws_break["A1"].font = font_section
    ws_break["G1"] = "SALES BY CUSTOMER SEGMENT"
    ws_break["G1"].font = font_section
    
    headers_cat = ["Category", "Revenue ($)", "Orders", "Avg Price ($)", "Gross Profit ($)", "Profit Margin %"]
    for i, h in enumerate(headers_cat):
        cell = ws_break.cell(row=3, column=i+1)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_accent
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    headers_seg = ["Age Group", "Revenue ($)", "Orders", "Unique Customers", "Revenue/Customer ($)"]
    for i, h in enumerate(headers_seg):
        cell = ws_break.cell(row=3, column=i+7)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_accent
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_break.row_dimensions[3].height = 22
    
    # Get Category breakdown
    cat_df = df.groupby('category').agg({
        'revenue': 'sum',
        'order_id': 'count',
        'unit_price': 'mean',
        'gross_profit': 'sum'
    }).sort_values('revenue', ascending=False)
    
    for idx, (cat_name, row_data) in enumerate(cat_df.iterrows()):
        r_idx = 4 + idx
        ws_break.row_dimensions[r_idx].height = 18
        
        c1 = ws_break.cell(row=r_idx, column=1, value=cat_name)
        c2 = ws_break.cell(row=r_idx, column=2, value=row_data['revenue'])
        c3 = ws_break.cell(row=r_idx, column=3, value=row_data['order_id'])
        c4 = ws_break.cell(row=r_idx, column=4, value=row_data['unit_price'])
        c5 = ws_break.cell(row=r_idx, column=5, value=row_data['gross_profit'])
        c6 = ws_break.cell(row=r_idx, column=6, value=row_data['gross_profit'] / row_data['revenue'])
        
        c2.number_format = "$#,##0.00"
        c3.number_format = "#,##0"
        c4.number_format = "$#,##0.00"
        c5.number_format = "$#,##0.00"
        c6.number_format = "0.00%"
        
        c1.alignment = Alignment(horizontal="left", vertical="center")
        for cell in [c2, c3, c4, c5, c6]:
            cell.alignment = Alignment(horizontal="right", vertical="center")
            
        for cell in [c1, c2, c3, c4, c5, c6]:
            cell.font = font_data
            cell.border = border_data
            if idx % 2 == 1:
                cell.fill = fill_zebra
                
    # Get Segment breakdown
    seg_df = df.groupby('age_group').agg({
        'revenue': 'sum',
        'order_id': 'count',
        'customer_id': 'nunique'
    }).sort_values('age_group')
    
    for idx, (seg_name, row_data) in enumerate(seg_df.iterrows()):
        r_idx = 4 + idx
        c1 = ws_break.cell(row=r_idx, column=7, value=seg_name)
        c2 = ws_break.cell(row=r_idx, column=8, value=row_data['revenue'])
        c3 = ws_break.cell(row=r_idx, column=9, value=row_data['order_id'])
        c4 = ws_break.cell(row=r_idx, column=10, value=row_data['customer_id'])
        c5 = ws_break.cell(row=r_idx, column=11, value=row_data['revenue'] / row_data['customer_id'])
        
        c2.number_format = "$#,##0.00"
        c3.number_format = "#,##0"
        c4.number_format = "#,##0"
        c5.number_format = "$#,##0.00"
        
        c1.alignment = Alignment(horizontal="center", vertical="center")
        for cell in [c2, c3, c4, c5]:
            cell.alignment = Alignment(horizontal="right", vertical="center")
            
        for cell in [c1, c2, c3, c4, c5]:
            cell.font = font_data
            cell.border = border_data
            if idx % 2 == 1:
                cell.fill = fill_zebra
                
    for col in ws_break.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 1 and cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws_break.column_dimensions[col_letter].width = max(max_len + 4, 15)
    ws_break.column_dimensions["A"].width = 18
    ws_break.column_dimensions["G"].width = 15
    ws_break.column_dimensions["F"].width = 5
    
    # Save output file
    wb.save('dashboard/sales_dashboard_mockup.xlsx')
    print("✨ Excel Mockup formatted and saved successfully!")

if __name__ == "__main__":
    build_beautiful_mockup()
